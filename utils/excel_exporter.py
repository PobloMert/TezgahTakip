from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime

class ExcelExporter:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
    def export_report(self, filename, data, report_type):
        # Başlık ekle
        self.ws.title = report_type[:30]  # Excel sheet isim sınırı
        
        # Üst bilgi
        self.ws['A1'] = f"Tezgah Takip Raporu - {report_type}"
        self.ws['A1'].font = Font(bold=True, size=14)
        self.ws.merge_cells('A1:D1')
        
        # Tarih bilgisi
        self.ws['A2'] = f"Oluşturulma Tarihi: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        self.ws.merge_cells('A2:D2')
        
        # Sütun başlıkları
        headers = ['Tezgah No', 'Bakım Sayısı', 'Son Bakım', 'Durum']
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        
        # Verileri ekle
        for row_num, row_data in enumerate(data, 4):
            self.ws.cell(row=row_num, column=1, value=row_data.get('machine_id', ''))
            self.ws.cell(row=row_num, column=2, value=row_data.get('maintenance_count', 0))
            self.ws.cell(row=row_num, column=3, value=row_data.get('last_maintenance', '-'))
            self.ws.cell(row=row_num, column=4, value=row_data.get('status', '-'))
        
        # Sütun genişliklerini ayarla
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            self.ws.column_dimensions[column].width = adjusted_width
        
        # Dosyayı kaydet
        self.wb.save(filename)
        return filename
